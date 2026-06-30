import os
import re
import sys
import csv
import time
import json
import traceback
from urllib.parse import quote

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# =========================
# CONFIG
# =========================
START_URL = "https://property.phila.gov/"
AIS_SEARCH_URL = "https://api.phila.gov/ais/v1/search"
OUTPUT_FILE_NAME = "Philadelphia_PA_AddressSearch_Tool_Output.xlsx"
INPUT_FILE_NAME = "PA.txt"

REQUEST_TIMEOUT = 45
SAVE_EVERY = 1

HEADERS = [
    "Sno",
    "County",
    "State",
    "Input",
    "Property Address",
    "Property CSZ",
    "Owner name",
    "Parcel number",
    "Mailing address",
    "Mailing address CSZ",
]


# =========================
# PATH HELPERS
# =========================
def get_run_folder():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


RUN_FOLDER = get_run_folder()
OUTPUT_PATH = os.path.join(RUN_FOLDER, OUTPUT_FILE_NAME)
INPUT_PATH = os.path.join(RUN_FOLDER, INPUT_FILE_NAME)


# =========================
# TEXT HELPERS
# =========================
def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_multiline(value):
    if value is None:
        return ""

    lines = []
    seen = set()

    for line in str(value).replace("\r", "\n").split("\n"):
        line = clean_text(line)
        if not line:
            continue

        key = line.upper()
        if key not in seen:
            lines.append(line)
            seen.add(key)

    return "\n".join(lines)


def is_opa_number(value):
    value = clean_text(value)
    return bool(re.fullmatch(r"\d{9}", value))


def format_csz_pipe_state(value):
    """
    PHILADELPHIA, PA 19134-2301 -> PHILADELPHIA, |PA| 19134-2301
    Philadelphia PA 19124       -> Philadelphia |PA| 19124
    """
    value = clean_text(value)
    if not value:
        return ""

    match = re.match(
        r"^(?P<city>.*?)(?P<comma>,?)\s+(?P<state>[A-Za-z]{2})\s+(?P<zip>\d{5}(?:-\d{4})?)$",
        value,
    )

    if not match:
        return value

    city = clean_text(match.group("city"))
    comma = match.group("comma")
    state = match.group("state").upper()
    zip_code = match.group("zip")

    if comma:
        return f"{city}, |{state}| {zip_code}"

    return f"{city} |{state}| {zip_code}"


def looks_like_csz(line):
    line = clean_text(line)
    return bool(
        re.match(
            r"^[A-Z][A-Z\s.'-]+,?\s+[A-Z]{2}\s+\d{5}(?:-\d{4})?$",
            line,
            flags=re.I,
        )
    )


def extract_csz_from_text(text):
    text = clean_text(text)
    if not text:
        return ""

    patterns = [
        r"PHILADELPHIA,\s*PA\s+\d{5}(?:-\d{4})?",
        r"PHILADELPHIA\s+PA\s+\d{5}(?:-\d{4})?",
        r"[A-Z][A-Z\s.'-]+,\s*PA\s+\d{5}(?:-\d{4})?",
        r"[A-Z][A-Z\s.'-]+\s+PA\s+\d{5}(?:-\d{4})?",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if match:
            return clean_text(match.group(0)).upper()

    return ""


def read_input_values(file_path):
    values = []
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xlsx":
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(values_only=True):
            for cell in row:
                text = clean_text(cell)
                if text:
                    values.append(text)
                    break

    elif ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                for cell in row:
                    text = clean_text(cell)
                    if text:
                        values.append(text)
                        break

    else:
        with open(file_path, "r", encoding="utf-8-sig") as f:
            for line in f:
                text = clean_text(line)
                if text:
                    values.append(text)

    final_values = []
    seen = set()

    for value in values:
        key = value.upper()
        if key not in seen:
            final_values.append(value)
            seen.add(key)

    return final_values


# =========================
# EXCEL HELPERS
# =========================
def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Philadelphia Property"
    ws.append(HEADERS)
    return wb, ws


def auto_fit_columns(ws):
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            for line in value.splitlines():
                max_len = max(max_len, len(line))

        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)


def save_workbook(wb, output_path):
    auto_fit_columns(wb.active)

    try:
        wb.save(output_path)
        return output_path

    except PermissionError:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        alt_path = os.path.join(
            RUN_FOLDER,
            f"Philadelphia_PA_AddressSearch_Tool_Output_{timestamp}.xlsx"
        )
        wb.save(alt_path)
        return alt_path


# =========================
# HTTP HELPERS
# =========================
def create_session():
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,application/json;q=0.8,*/*;q=0.7",
            "Accept-Language": "en-US,en;q=0.9",
        }
    )
    return session


def get_json(session, url):
    response = session.get(url, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    return response.json()


def get_html(session, url):
    response = session.get(url, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    return response.text


# =========================
# AIS API HELPERS
# =========================
def collect_opa_numbers(obj, results=None):
    if results is None:
        results = []

    if isinstance(obj, dict):
        for key, value in obj.items():
            key_text = str(key).lower()

            if any(word in key_text for word in ["opa", "account", "brt"]):
                collect_opa_numbers(value, results)

            # Also continue recursive search because AIS response structures can change.
            if isinstance(value, (dict, list)):
                collect_opa_numbers(value, results)

    elif isinstance(obj, list):
        for item in obj:
            collect_opa_numbers(item, results)

    else:
        text = clean_text(obj)
        for match in re.findall(r"\b\d{9}\b", text):
            if match not in results:
                results.append(match)

    return results


def find_first_value_by_keys(obj, key_candidates):
    key_candidates = [k.lower() for k in key_candidates]

    if isinstance(obj, dict):
        for key, value in obj.items():
            key_text = str(key).lower()

            if key_text in key_candidates:
                result = clean_text(value)
                if result:
                    return result

            if any(candidate in key_text for candidate in key_candidates):
                result = clean_text(value)
                if result and not isinstance(value, (dict, list)):
                    return result

        for value in obj.values():
            result = find_first_value_by_keys(value, key_candidates)
            if result:
                return result

    elif isinstance(obj, list):
        for item in obj:
            result = find_first_value_by_keys(item, key_candidates)
            if result:
                return result

    return ""


def get_opa_from_ais(session, input_value):
    """
    Uses Philadelphia AIS /search endpoint to convert an address to an OPA account number.
    If the input is already a 9-digit OPA number, it returns that directly.
    """
    input_value = clean_text(input_value)

    if is_opa_number(input_value):
        return {
            "opa_number": input_value,
            "ais_address": "",
            "ais_csz": "",
            "ais_raw": {},
        }

    url = f"{AIS_SEARCH_URL}/{quote(input_value)}?opa_only"

    # Optional API key support, if the City requires one later.
    gatekeeper_key = os.environ.get("AIS_GATEKEEPER_KEY", "").strip()
    if gatekeeper_key:
        url += f"&gatekeeperKey={quote(gatekeeper_key)}"

    print(f"AIS search URL: {url}")

    data = get_json(session, url)
    features = data.get("features", []) if isinstance(data, dict) else []

    if not features:
        return {
            "opa_number": "",
            "ais_address": "",
            "ais_csz": "",
            "ais_raw": data,
        }

    first_feature = features[0]
    props = first_feature.get("properties", {}) if isinstance(first_feature, dict) else {}

    opa_numbers = collect_opa_numbers(first_feature)
    opa_number = opa_numbers[0] if opa_numbers else ""

    ais_address = find_first_value_by_keys(
        props,
        [
            "street_address",
            "address",
            "full_address",
            "standardized_address",
            "opa_address",
        ],
    )

    ais_csz = extract_csz_from_text(json.dumps(props))

    return {
        "opa_number": opa_number,
        "ais_address": ais_address,
        "ais_csz": ais_csz,
        "ais_raw": data,
    }


# =========================
# PROPERTY PAGE PARSING
# =========================
STOP_LABELS = {
    "PROPERTY ASSESSMENT AND SALE INFORMATION",
    "ASSESSED VALUE",
    "SALE DATE",
    "SALE PRICE",
    "PROPERTY CHARACTERISTICS",
    "EXEMPTION INFORMATION",
    "TAX BALANCES",
    "VIEW TAX BALANCE",
    "SUBMIT AN OFFICIAL INQUIRY",
    "PRINT",
}


def html_to_lines(html):
    soup = BeautifulSoup(html, "html.parser")

    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()

    text = soup.get_text("\n")
    raw_lines = [clean_text(line) for line in text.splitlines()]

    lines = []
    for line in raw_lines:
        if not line:
            continue
        if line not in lines:
            lines.append(line)

    return lines


def find_line_index(lines, text, start=0):
    wanted = clean_text(text).lower()
    for i in range(start, len(lines)):
        if clean_text(lines[i]).lower() == wanted:
            return i
    return -1


def extract_property_address_and_csz(lines, ais_address="", ais_csz=""):
    property_address = clean_text(ais_address).upper()
    property_csz = clean_text(ais_csz).upper()

    # First CSZ near top is normally the property CSZ.
    for i, line in enumerate(lines[:80]):
        if looks_like_csz(line):
            property_csz = clean_text(line).upper()

            # Previous non-label line is normally property address.
            for j in range(i - 1, -1, -1):
                candidate = clean_text(lines[j])
                if not candidate:
                    continue

                if candidate.upper() in STOP_LABELS:
                    continue

                if candidate.lower() in {"property", "print"}:
                    continue

                if looks_like_csz(candidate):
                    continue

                property_address = candidate.upper()
                break

            break

    return property_address, property_csz


def extract_owner_and_mailing(lines, opa_number):
    owner_name = ""
    parcel_number = clean_text(opa_number)
    mailing_address = ""
    mailing_csz = ""

    opa_idx = -1
    for i, line in enumerate(lines):
        if clean_text(line) == parcel_number:
            opa_idx = i
            break

    if opa_idx == -1:
        # If exact OPA line is not found, use first 9-digit number.
        for i, line in enumerate(lines):
            if is_opa_number(line):
                parcel_number = clean_text(line)
                opa_idx = i
                break

    # Owner is usually between "OPA Account Number" label and the OPA number.
    if opa_idx != -1:
        account_label_idx = -1
        for i in range(0, opa_idx):
            if "opa account number" in clean_text(lines[i]).lower():
                account_label_idx = i

        owner_start = account_label_idx + 1 if account_label_idx != -1 else max(0, opa_idx - 4)
        owner_lines = []

        for line in lines[owner_start:opa_idx]:
            value = clean_text(line)
            if not value:
                continue
            if value.upper() in STOP_LABELS:
                continue
            if value.lower() in {"owner", "opa account number", "property"}:
                continue
            if looks_like_csz(value):
                continue
            if is_opa_number(value):
                continue
            owner_lines.append(value)

        owner_name = clean_multiline("\n".join(owner_lines))

    # Mailing address block
    mailing_idx = find_line_index(lines, "Mailing Address", start=max(0, opa_idx))
    if mailing_idx == -1:
        mailing_idx = find_line_index(lines, "Mailing address", start=0)

    if mailing_idx != -1:
        mail_lines = []

        for line in lines[mailing_idx + 1:]:
            value = clean_text(line)
            upper = value.upper()

            if not value:
                continue

            if upper in STOP_LABELS:
                break

            if "PROPERTY ASSESSMENT" in upper:
                break

            if value.lower() in {"owner", "opa account number", "mailing address"}:
                break

            if is_opa_number(value):
                break

            mail_lines.append(value)

            # Once we have a CSZ line, stop after that.
            if looks_like_csz(value) and len(mail_lines) >= 2:
                break

        if mail_lines:
            if looks_like_csz(mail_lines[-1]):
                mailing_csz = mail_lines[-1]
                mailing_address = "\n".join(mail_lines[:-1])
            else:
                mailing_address = "\n".join(mail_lines)

    return {
        "owner_name": owner_name,
        "parcel_number": parcel_number,
        "mailing_address": clean_multiline(mailing_address),
        "mailing_csz": clean_text(mailing_csz),
    }


def scrape_property_page(session, opa_number, ais_address="", ais_csz=""):
    url = f"{START_URL}?p={quote(clean_text(opa_number))}"
    print(f"Property page URL: {url}")

    html = get_html(session, url)
    lines = html_to_lines(html)

    property_address, property_csz = extract_property_address_and_csz(
        lines,
        ais_address=ais_address,
        ais_csz=ais_csz,
    )

    owner_data = extract_owner_and_mailing(lines, opa_number)

    return {
        "opa_address": property_address,
        "property_csz": property_csz,
        "owner_name": owner_data.get("owner_name", ""),
        "parcel_number": owner_data.get("parcel_number", opa_number),
        "mailing_address": owner_data.get("mailing_address", ""),
        "mailing_csz": owner_data.get("mailing_csz", ""),
    }


# =========================
# ROW HELPERS
# =========================
def make_output_row(serial_no, input_value, details):
    property_address = clean_multiline(details.get("opa_address", ""))
    if property_address:
        property_address = property_address.upper()

    property_csz = format_csz_pipe_state(details.get("property_csz", ""))
    mailing_address = clean_multiline(details.get("mailing_address", ""))
    mailing_csz = format_csz_pipe_state(details.get("mailing_csz", ""))

    return [
        serial_no,
        "Philadelphia",
        "PA",
        input_value,
        property_address,
        property_csz,
        clean_multiline(details.get("owner_name", "")),
        clean_text(details.get("parcel_number", "")),
        mailing_address,
        mailing_csz,
    ]


def make_blank_row(serial_no, input_value, remarks="No Record"):
    return [
        serial_no,
        "Philadelphia",
        "PA",
        input_value,
        "",
        "",
        "",
        remarks,
        "",
        "",
    ]


def scrape_one_input(session, input_value):
    ais_result = get_opa_from_ais(session, input_value)

    opa_number = clean_text(ais_result.get("opa_number", ""))

    if not opa_number:
        raise Exception("OPA account number not found from AIS search")

    details = scrape_property_page(
        session,
        opa_number,
        ais_address=ais_result.get("ais_address", ""),
        ais_csz=ais_result.get("ais_csz", ""),
    )

    if not details.get("parcel_number"):
        details["parcel_number"] = opa_number

    return details


# =========================
# MAIN
# =========================
def main():
    final_output_path = OUTPUT_PATH

    try:
        input_file = INPUT_PATH

        if not os.path.exists(input_file):
            print(f"ERROR: Input file not found: {input_file}")
            print("Create PA.txt in the same folder as this script and run again.")
            return

        input_values = read_input_values(input_file)

        if not input_values:
            print(f"ERROR: Input file is empty: {input_file}")
            return

        session = create_session()
        wb, ws = create_workbook()

        for index, input_value in enumerate(input_values, start=1):
            print(f"[{index}/{len(input_values)}] Searching: {input_value}")

            try:
                details = scrape_one_input(session, input_value)
                row = make_output_row(index, input_value, details)
                ws.append(row)

                print(f"Done: {input_value}")

            except Exception as e:
                print(f"Error for {input_value}: {e}")
                traceback.print_exc()
                ws.append(make_blank_row(index, input_value, "No Record"))

            if index % SAVE_EVERY == 0:
                final_output_path = save_workbook(wb, final_output_path)

        final_output_path = save_workbook(wb, final_output_path)

        print("All records completed.")
        print(f"Excel saved here: {final_output_path}")

    except Exception as e:
        traceback.print_exc()
        print(f"ERROR: Tool stopped because of error: {e}")


if __name__ == "__main__":
    main()
